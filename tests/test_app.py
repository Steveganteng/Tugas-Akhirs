"""Pengujian fungsional (black-box) aplikasi web restok obat.

Setiap fungsi tes berkorespondensi dengan satu baris di docs/TESTCASE.md (kode TC).
Jalankan: py -3 -m pytest -v
"""
import io
import re
from openpyxl import load_workbook


# ----------------------- AUTENTIKASI -----------------------
def test_TC01_login_page_render(client):
    """TC-01a: Halaman login tampil dengan form."""
    r = client.get("/login")
    html = r.get_data(as_text=True)
    assert r.status_code == 200
    assert 'name="username"' in html and 'name="password"' in html


def test_TC02_proteksi_tanpa_login(client):
    """TC-02: Akses halaman tanpa login diarahkan ke /login."""
    r = client.get("/")
    assert r.status_code == 302
    assert "/login" in r.headers.get("Location", "")


def test_TC03_login_berhasil(client):
    """TC-03: Login benar -> diarahkan ke dashboard."""
    r = client.post("/login", data={"username": "apoteker", "password": "apoteker123"})
    assert r.status_code == 302
    assert r.headers.get("Location", "").endswith("/")
    # halaman utama bisa diakses sesudahnya
    assert client.get("/").status_code == 200


def test_TC04_login_gagal(client):
    """TC-04: Login salah -> pesan kesalahan, tetap di halaman login."""
    r = client.post("/login", data={"username": "apoteker", "password": "salah"})
    assert r.status_code == 200
    assert "salah" in r.get_data(as_text=True).lower()


def test_TC05_logout(auth):
    """TC-05: Logout -> sesi berakhir, halaman terproteksi lagi."""
    r = auth.get("/logout")
    assert r.status_code == 302 and "/login" in r.headers.get("Location", "")
    assert auth.get("/").status_code == 302  # sudah tidak bisa akses


# ----------------------- DASHBOARD -----------------------
def test_TC06_dashboard_tampil(auth):
    """TC-06: Dashboard menampilkan ringkasan & tabel butuh restok."""
    html = auth.get("/").get_data(as_text=True)
    assert "Butuh Segera Direstok" in html
    assert "Total Obat" in html


def _table_headers(html, after_marker):
    seg = html.split(after_marker, 1)[1]
    head = seg.split("</thead>", 1)[0]
    ths = re.findall(r"<th[^>]*>(.*?)</th>", head, re.S)
    return [re.sub(r"<[^>]+>", "", t).strip() for t in ths if re.sub(r"<[^>]+>", "", t).strip()]


def test_TC07_header_dashboard_selaras_rekomendasi(auth):
    """TC-07: Header tabel dashboard sama dengan tabel rekomendasi."""
    dash = _table_headers(auth.get("/").get_data(as_text=True), "Butuh Segera Direstok")
    rek = _table_headers(auth.get("/rekomendasi/").get_data(as_text=True), "ditemukan")
    assert dash == rek, f"dashboard={dash} != rekomendasi={rek}"


# ----------------------- REKOMENDASI -----------------------
def _badge_counts(html):
    return {c: len(re.findall(c, html)) for c in ["s-segera", "s-perhatian", "s-aman"]}


def test_TC08_filter_status_segera(auth):
    """TC-08: Filter status SEGERA RESTOK -> hanya badge segera."""
    html = auth.get("/rekomendasi/?status=SEGERA+RESTOK").get_data(as_text=True)
    c = _badge_counts(html)
    assert c["s-segera"] > 0
    assert c["s-aman"] == 0 and c["s-perhatian"] == 0


def test_TC09_status_kosong_tidak_mengosongkan(auth):
    """TC-09: 'Semua status' (kosong) tidak menyaring habis semua baris."""
    html = auth.get("/rekomendasi/?status=").get_data(as_text=True)
    m = re.search(r"(\d+)\s+obat ditemukan", html)
    assert m and int(m.group(1)) > 0


def test_TC10_periode_historis(auth):
    """TC-10: Periode lampau menampilkan data aktual (status HISTORIS)."""
    html = auth.get("/rekomendasi/?periode=2025-06").get_data(as_text=True)
    assert "HISTORIS" in html


def test_TC11_sorting(auth):
    """TC-11: Sorting kolom dapat diterapkan tanpa error."""
    assert auth.get("/rekomendasi/?sort=stok_saat_ini&dir=asc").status_code == 200


def test_TC12_ekspor_excel(auth):
    """TC-12: Ekspor Excel menghasilkan .xlsx dengan header yang benar."""
    r = auth.get("/rekomendasi/export")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("Content-Type", "")
    ws = load_workbook(io.BytesIO(r.get_data())).active
    header = [c.value for c in ws[1]]
    for col in ["Nama Obat", "Jenis Data", "Satuan", "Status"]:
        assert col in header


# ----------------------- PREDIKSI -----------------------
def test_TC13_prediksi_satu_obat(auth):
    """TC-13: Prediksi satu obat menampilkan satuan & tabel forecast."""
    r = auth.get("/prediksi/?obat=ACETYLCYSTEIN+KAP&horizon=6")
    html = r.get_data(as_text=True)
    assert r.status_code == 200
    assert "Satuan" in html and "KAPSUL" in html


def test_TC14_prediksi_total(auth):
    """TC-14: Mode TOTAL semua obat dapat dibuka."""
    assert auth.get("/prediksi/?obat=__TOTAL__&horizon=12").status_code == 200


# ----------------------- UPLOAD -----------------------
def test_TC15_unduh_template(auth):
    """TC-15: Template upload dapat diunduh (.xlsx, kolom wajib)."""
    r = auth.get("/upload/template")
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.get_data())).active
    header = [c.value for c in ws[1]]
    for col in ["Tanggal Masuk", "Resep Obat", "JUMLAH", "SISA_STOK"]:
        assert col in header


def test_TC16_validasi_file_invalid(app):
    """TC-16: File tanpa kolom wajib ditolak validator."""
    import pandas as pd
    from app.services.upload_service import validate_dataframe
    df = pd.DataFrame({"kolom_asal": [1, 2, 3]})
    valid, n, nv, nr, errors = validate_dataframe(df)
    assert valid is None and errors


# ----------------------- MODEL -----------------------
def test_TC17_model_ses_terdaftar(app):
    """TC-17: Model SES terdaftar & dapat dimuat; tepat satu model aktif."""
    with app.app_context():
        from app.models import ModelRegistry
        from app.services.forecast_service import service as FS
        assert ModelRegistry.query.filter_by(status="active").count() == 1
        assert ModelRegistry.query.filter_by(nama_model="ses").first() is not None
        FS.load_active()
        assert "ses" in FS.models  # SES tersedia sebagai opsi model
