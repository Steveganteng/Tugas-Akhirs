"""Upload data pengeluaran: template, validasi, ingest ke transaksi_raw.

Kolom raw WAJIB (PERSIS):
  Tanggal Masuk | Register | Kode Diagnosa | Diagnosa Primer | Resep Obat |
  JUMLAH | SISA_STOK | SATUAN
Catatan: JUMLAH & SISA_STOK adalah agregat bulanan per obat (broadcast), bukan
nilai per transaksi -> JANGAN dijumlah saat membentuk panel.
"""
from __future__ import annotations
import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import config as C
from data_processing import clean, build_panel, load_raw
from ..extensions import db
from ..models import TransaksiRaw, UploadLog

RAW_COLS = ["Tanggal Masuk", "Register", "Kode Diagnosa", "Diagnosa Primer",
            "Resep Obat", "JUMLAH", "SISA_STOK", "SATUAN"]
REQUIRED = ["Tanggal Masuk", "Resep Obat", "JUMLAH", "SISA_STOK"]
TOLERANSI_TOLAK = 0.20  # > 20% baris tak valid -> upload gagal


# ---------------- TEMPLATE ----------------
def generate_template() -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Kunjungan"
    head_fill = PatternFill("solid", fgColor="2E7D32")
    head_font = Font(color="FFFFFF", bold=True)
    for j, col in enumerate(RAW_COLS, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.fill = head_fill
        c.font = head_font
        ws.column_dimensions[c.column_letter].width = max(14, len(col) + 3)
    contoh = [
        ["2025-01-05", "RJ0001", "J06.9", "ISPA", "PARACETAMOL TABLET", 1200, 3400, "TABLET"],
        ["2025-01-12", "RJ0002", "K30",   "Dispepsia", "ANTASIDA TABLET", 850, 2100, "TABLET"],
        ["2025-02-03", "RJ0003", "J06.9", "ISPA", "PARACETAMOL TABLET", 1100, 2900, "TABLET"],
    ]
    for i, row in enumerate(contoh, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    pet = wb.create_sheet("PETUNJUK")
    pet["A1"] = "PETUNJUK PENGISIAN TEMPLATE"
    pet["A1"].font = Font(bold=True, size=13)
    petunjuk = [
        ("Kolom", "Format / Keterangan"),
        ("Tanggal Masuk", "Tanggal (YYYY-MM-DD). WAJIB. Dipakai menentukan bulan."),
        ("Register", "Teks bebas nomor register kunjungan. Opsional."),
        ("Kode Diagnosa", "Kode ICD. Opsional."),
        ("Diagnosa Primer", "Nama diagnosa. Opsional."),
        ("Resep Obat", "Nama obat. WAJIB. Akan dinormalisasi huruf besar."),
        ("JUMLAH", "Angka. WAJIB. = total pengeluaran obat dalam BULAN tsb."),
        ("SISA_STOK", "Angka. WAJIB. = sisa stok obat pada BULAN tsb."),
        ("SATUAN", "Satuan obat (TABLET/KAP/BTL/...). Opsional."),
        ("", ""),
        ("PENTING", "JUMLAH & SISA_STOK adalah AGREGAT BULANAN per obat yang"),
        ("", "di-broadcast ke tiap baris. Nilai sama untuk obat-bulan yang sama."),
        ("", "Sistem mengambil first() non-null per (obat, bulan), BUKAN menjumlah."),
    ]
    for i, (a, b) in enumerate(petunjuk, start=3):
        ca = pet.cell(row=i, column=1, value=a)
        pet.cell(row=i, column=2, value=b)
        if a in ("Kolom", "PENTING"):
            ca.font = Font(bold=True)
    pet.column_dimensions["A"].width = 18
    pet.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------- VALIDASI ----------------
def validate_dataframe(df: pd.DataFrame):
    """Kembalikan (df_valid_raw, n_baris, n_valid, n_ditolak, errors:list[str])."""
    errors = []
    missing = [c for c in REQUIRED if c not in df.columns]
    if missing:
        errors.append(f"Kolom wajib hilang: {', '.join(missing)}")
        return None, len(df), 0, len(df), errors

    n_baris = len(df)
    work = df.copy()
    work["__tgl"] = pd.to_datetime(work["Tanggal Masuk"], errors="coerce")
    work["__jml"] = pd.to_numeric(work["JUMLAH"], errors="coerce")
    work["__stok"] = pd.to_numeric(work["SISA_STOK"], errors="coerce")
    work["__obat"] = work["Resep Obat"].astype("string").str.strip()

    valid_mask = (work["__tgl"].notna() & work["__jml"].notna()
                  & work["__stok"].notna() & work["__obat"].notna()
                  & (work["__obat"] != ""))
    n_valid = int(valid_mask.sum())
    n_ditolak = n_baris - n_valid
    if n_baris == 0:
        errors.append("File kosong / tidak ada baris data.")
        return None, 0, 0, 0, errors
    if n_ditolak / n_baris > TOLERANSI_TOLAK:
        errors.append(f"{n_ditolak}/{n_baris} baris tidak valid "
                      f"(> {int(TOLERANSI_TOLAK*100)}%). Upload ditolak.")
        return None, n_baris, n_valid, n_ditolak, errors

    valid = df[valid_mask.values].copy()
    # pastikan kolom raw lengkap (isi yang opsional bila tak ada)
    for c in RAW_COLS:
        if c not in valid.columns:
            valid[c] = None
    return valid[RAW_COLS], n_baris, n_valid, n_ditolak, errors


# ---------------- INGEST ----------------
def stage_upload(filename: str, df_valid: pd.DataFrame, n_baris, n_valid,
                 n_ditolak, model_pilihan: str) -> int:
    """Buat baris upload_log + simpan transaksi_raw. Kembalikan upload_id."""
    log = UploadLog(filename=filename, n_baris=n_baris, n_valid=n_valid,
                    n_ditolak=n_ditolak, status="diproses",
                    model_dilatih=model_pilihan)
    db.session.add(log)
    db.session.flush()  # dapatkan upload_id
    uid = log.upload_id

    for _, r in df_valid.iterrows():
        tgl = pd.to_datetime(r["Tanggal Masuk"], errors="coerce")
        db.session.add(TransaksiRaw(
            tanggal_masuk=(tgl.date() if pd.notna(tgl) else None),
            register=_s(r.get("Register")),
            kode_diagnosa=_s(r.get("Kode Diagnosa")),
            diagnosa_primer=_s(r.get("Diagnosa Primer")),
            resep_obat=_s(r.get("Resep Obat")),
            jumlah=_f(r.get("JUMLAH")),
            sisa_stok=_f(r.get("SISA_STOK")),
            satuan=_s(r.get("SATUAN")),
            upload_id=uid,
        ))
    db.session.commit()
    return uid


def raw_from_db() -> pd.DataFrame:
    """Ambil SEMUA transaksi_raw (seluruh riwayat upload) sebagai DataFrame
    dengan nama kolom raw asli."""
    rows = TransaksiRaw.query.all()
    recs = []
    for r in rows:
        recs.append({
            "Tanggal Masuk": r.tanggal_masuk, "Register": r.register,
            "Kode Diagnosa": r.kode_diagnosa, "Diagnosa Primer": r.diagnosa_primer,
            "Resep Obat": r.resep_obat, "JUMLAH": r.jumlah,
            "SISA_STOK": r.sisa_stok, "SATUAN": r.satuan,
        })
    return pd.DataFrame(recs, columns=RAW_COLS)


def build_combined_panel(df_new_raw: pd.DataFrame | None = None) -> pd.DataFrame:
    """Preprocessing penuh untuk challenger.

    Gabungkan: raw asli (sumber Excel) + SELURUH transaksi_raw (semua upload
    sebelumnya + upload sekarang yang sudah di-stage) -> clean() -> build_panel().
    clean(): parse tanggal, normalisasi nama obat (strip+upper), buang baris tak
    valid, dedup. build_panel(): first() non-null per (obat,bulan) (anti
    double-count), reindex 12 bulan, isi celah. Hasil: panel obat×bulan siap latih.
    """
    frames = []
    try:
        frames.append(load_raw())
    except Exception:
        pass
    db_raw = raw_from_db()
    if len(db_raw):
        frames.append(db_raw)
    if df_new_raw is not None and len(df_new_raw):
        frames.append(df_new_raw)  # jaring pengaman bila belum ter-stage
    combined = (pd.concat(frames, ignore_index=True)
                if frames else pd.DataFrame(columns=RAW_COLS))
    cleaned = clean(combined)            # <-- PREPROCESSING tahap 1
    panel = build_panel(cleaned)          # <-- PREPROCESSING tahap 2
    return panel


def _s(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    return str(v).strip()


def _f(v):
    try:
        f = float(v)
        return f
    except (TypeError, ValueError):
        return None
